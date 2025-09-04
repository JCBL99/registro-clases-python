# app_streamlit.py
import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

# Configuración de la página
st.set_page_config(
    page_title="Sistema de Registro de Clases",
    page_icon="📚",
    layout="wide"
)

# Título de la aplicación
st.title("📚 Sistema de Registro de Clases")
st.markdown("---")

class RegistroClases:
    def __init__(self, archivo_excel="registro_clases.xlsx"):
        self.archivo_excel = archivo_excel
        self.inicializar_archivo()
    
    def inicializar_archivo(self):
        if not os.path.exists(self.archivo_excel):
            wb = Workbook()
            ws = wb.active
            ws.title = "Registro de Clases"
            
            encabezados = ["Nombre de la Clase", "Fecha", "Duración (minutos)", "Costo", "Total"]
            for col, encabezado in enumerate(encabezados, 1):
                ws.cell(row=1, column=col, value=encabezado)
            
            ws.cell(row=2, column=5, value="=SUM(D2:D1000)")
            wb.save(self.archivo_excel)
    
    def calcular_costo(self, minutos):
        if minutos == 90:
            return 32250
        elif minutos == 80:
            return 28667
        elif minutos == 60:
            return 21500
        else:
            costo_por_minuto = 21500 / 60
            return round(minutos * costo_por_minuto)
    
    def agregar_clase(self, nombre_clase, fecha, minutos):
        try:
            wb = load_workbook(self.archivo_excel)
            ws = wb.active
            
            siguiente_fila = ws.max_row + 1
            if siguiente_fila < 3:
                siguiente_fila = 3
            
            costo = self.calcular_costo(minutos)
            
            ws.cell(row=siguiente_fila, column=1, value=nombre_clase)
            ws.cell(row=siguiente_fila, column=2, value=fecha)
            ws.cell(row=siguiente_fila, column=3, value=minutos)
            ws.cell(row=siguiente_fila, column=4, value=costo)
            
            wb.save(self.archivo_excel)
            return True, costo
            
        except Exception as e:
            return False, str(e)
    
    def obtener_clases(self):
        try:
            wb = load_workbook(self.archivo_excel)
            ws = wb.active
            
            clases = []
            total_general = 0
            
            for fila, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 1):
                if row[0] and row[0] != "Total General":
                    clase = {
                        "No.": fila,
                        "Nombre": row[0],
                        "Fecha": row[1],
                        "Duración (min)": row[2],
                        "Costo ($)": row[3]
                    }
                    clases.append(clase)
                    total_general += row[3] if row[3] else 0
            
            return clases, total_general
            
        except Exception as e:
            return [], 0
    
    def eliminar_clase(self, numero_fila):
        try:
            wb = load_workbook(self.archivo_excel)
            ws = wb.active
            
            fila_real = numero_fila + 2
            if fila_real > ws.max_row or fila_real < 3:
                return False
            
            clase_eliminada = ws.cell(row=fila_real, column=1).value
            ws.delete_rows(fila_real)
            wb.save(self.archivo_excel)
            
            return True, clase_eliminada
            
        except Exception as e:
            return False, str(e)

# Inicializar la aplicación
registro = RegistroClases()

# Sidebar para agregar clases
with st.sidebar:
    st.header("➕ Agregar Nueva Clase")
    
    with st.form("agregar_clase"):
        nombre = st.text_input("Nombre de la clase:")
        fecha = st.date_input("Fecha:")
        duracion = st.selectbox("Duración (minutos):", [60, 80, 90])
        
        submitted = st.form_submit_button("Agregar Clase")
        
        if submitted:
            if nombre and fecha:
                fecha_str = fecha.strftime("%Y-%m-%d")
                success, resultado = registro.agregar_clase(nombre, fecha_str, duracion)
                
                if success:
                    st.success(f"✅ Clase '{nombre}' agregada! Costo: ${resultado}")
                    st.rerun()
                else:
                    st.error(f"❌ Error: {resultado}")
            else:
                st.warning("⚠️ Por favor complete todos los campos")

# Mostrar clases registradas
st.header("📋 Clases Registradas")

clases, total_general = registro.obtener_clases()

if clases:
    # Convertir a DataFrame para mejor visualización
    df = pd.DataFrame(clases)
    df.set_index("No.", inplace=True)
    
    # Mostrar tabla
    st.dataframe(
        df,
        use_container_width=True,
        height=400
    )
    
    # Mostrar total general
    st.metric("💰 Total General", f"${total_general:,.0f}")
    
    # Opción para eliminar
    st.subheader("🗑️ Eliminar Clase")
    col1, col2 = st.columns([2, 1])
    
    with col1:
        clase_a_eliminar = st.selectbox(
            "Seleccione la clase a eliminar:",
            options=[f"{clase['No.']} - {clase['Nombre']} - {clase['Fecha']}" for clase in clases],
            key="eliminar_select"
        )
    
    with col2:
        if st.button("Eliminar Seleccionada", type="secondary"):
            if clase_a_eliminar:
                numero_clase = int(clase_a_eliminar.split(" - ")[0])
                success, resultado = registro.eliminar_clase(numero_clase)
                
                if success:
                    st.success(f"✅ Clase '{resultado}' eliminada!")
                    st.rerun()
                else:
                    st.error(f"❌ Error: {resultado}")
            else:
                st.warning("⚠️ Seleccione una clase para eliminar")

else:
    st.info("📝 No hay clases registradas. Agrega la primera clase usando el formulario a la izquierda.")

# Información del archivo
st.sidebar.markdown("---")
st.sidebar.info(f"📊 **Datos guardados en:** {registro.archivo_excel}")
st.sidebar.download_button(
    label="📥 Descargar Excel",
    data=open(registro.archivo_excel, "rb").read(),
    file_name=registro.archivo_excel,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

if __name__ == "__main__":
    pass